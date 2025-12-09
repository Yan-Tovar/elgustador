from rest_framework.views import APIView
from rest_framework import viewsets, status, filters
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from config.permissions import IsAdministrador
from django.contrib.auth import authenticate
from .models import Usuario
from .serializers import (
    UsuarioSerializer,
    UsuarioRegistroSerializer,
    UsuarioLoginSerializer,
    UsuarioUpdateSerializer,
    UsuarioRolUpdateSerializer
)
from rest_framework_simplejwt.tokens import RefreshToken
from notificaciones.services import crear_notificacion

from .paginators import UsuarioPagination


from django.db.models import Q
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count
from django.db.models.functions import TruncMonth
from openpyxl.styles import Font
from django.utils.timezone import now


def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)
    return {
        "refresh": str(refresh),
        "access": str(refresh.access_token),
    }


class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all().select_related("departamento", "municipio")
    serializer_class = UsuarioSerializer
    permission_classes = [IsAuthenticated]


class UsuarioRegistroViewSet(viewsets.GenericViewSet):
    serializer_class = UsuarioRegistroSerializer
    permission_classes = [AllowAny]

    def create(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        usuario = serializer.save()

        # =========================
        # Crear notificación de bienvenida
        # =========================
        crear_notificacion(
            usuario=usuario,
            titulo="¡Bienvenido a El Gustador!",
            mensaje="Te damos la bienvenida al sistema. Esperamos que tengas una excelente experiencia.",
            enviar_email=True  
        )

        return Response(UsuarioSerializer(usuario).data, status=status.HTTP_201_CREATED)


class LoginViewSet(viewsets.GenericViewSet):
    serializer_class = UsuarioLoginSerializer
    permission_classes = [AllowAny]

    def create(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        identificacion = serializer.validated_data["identificacion"]
        password = serializer.validated_data["password"]
        user = authenticate(request, identificacion=identificacion, password=password)
        if not user:
            return Response({"detail": "Credenciales incorrectas"}, status=400)

        tokens = get_tokens_for_user(user)
        return Response({
            "user": UsuarioSerializer(user).data,
            **tokens
        })


class UsuarioSelfViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    # GET /usuarios/usuarios/me/
    def list(self, request):
        usuario = request.user
        serializer = UsuarioSerializer(usuario)
        return Response(serializer.data)

    # PUT /usuarios/usuarios/me/
    def update(self, request):
        usuario = request.user
        serializer = UsuarioUpdateSerializer(
            usuario, data=request.data, partial=True
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # Devolver datos completos con relaciones anidadas
        return Response(UsuarioSerializer(usuario).data, status=status.HTTP_200_OK)


class UsuarioAdminViewSet(viewsets.ModelViewSet):
    """
    ADMIN:
    - Listar usuarios (paginado)
    - Buscar por texto (id, email, rol)
    - Filtrar por rol (botones)
    - Ver detalle de usuario
    - Cambiar rol
    - Envía notificación y correo al cambiar rol
    """

    serializer_class = UsuarioSerializer
    permission_classes = [IsAuthenticated, IsAdministrador]
    pagination_class = UsuarioPagination

    filter_backends = [filters.SearchFilter]
    search_fields = [
        "identificacion",
        "email",
        "rol",
    ]

    def get_queryset(self):
        queryset = (
            Usuario.objects
            .all()
            .select_related("departamento", "municipio")
            .order_by("-id")
        )

        # 🔹 Filtro por rol (desde query param)
        rol = self.request.query_params.get("rol")
        if rol:
            queryset = queryset.filter(rol=rol)

        return queryset

    def partial_update(self, request, *args, **kwargs):
        """
        Solo ADMIN puede cambiar rol
        Envía notificación/correo según el cambio
        """
        usuario = self.get_object()
        rol_anterior = usuario.rol

        serializer = UsuarioRolUpdateSerializer(
            usuario,
            data=request.data,
            partial=True
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        rol_nuevo = serializer.validated_data.get("rol", rol_anterior)

        #  Si cambió el rol, enviamos notificación y correo
        if rol_anterior != rol_nuevo:
            if rol_anterior == "cliente" and rol_nuevo in ["empleado", "admin"]:
                # Bienvenida a equipo
                crear_notificacion(
                    usuario=usuario,
                    titulo="¡Bienvenido al equipo de Aliños El Gustador!",
                    mensaje=(
                        f"Hola {usuario.nombre}, ahora eres parte del equipo como {rol_nuevo}. "
                        "Esperamos que disfrutes tu experiencia y aportes mucho al Gustador."
                    ),
                    enviar_email=True
                )
            elif rol_anterior in ["empleado", "admin"] and rol_nuevo == "cliente":
                # Agradecimiento por dejar el equipo
                crear_notificacion(
                    usuario=usuario,
                    titulo="Gracias por haber sido parte del equipo de El Gustador",
                    mensaje=(
                        f"Hola {usuario.nombre}, gracias por tu tiempo como {rol_anterior}. "
                        "Ahora eres cliente, y esperamos que sigas disfrutando nuestros productos."
                    ),
                    enviar_email=True
                )

        return Response(
            UsuarioSerializer(usuario).data,
            status=status.HTTP_200_OK
        )

# =============================
# ESTADÍSTICA: USUARIOS POR ROL
# =============================
class UsuariosPorRolView(APIView):
    permission_classes = [IsAuthenticated, IsAdministrador]

    def get(self, request):
        data = (
            Usuario.objects
            .values("rol")
            .annotate(total=Count("id"))
            .order_by("rol")
        )
        return Response(data)


# =============================
# ESTADÍSTICA: USUARIOS POR MES
# =============================
class UsuariosPorMesView(APIView):
    permission_classes = [IsAuthenticated, IsAdministrador]

    def get(self, request):
        data = (
            Usuario.objects
            .annotate(mes=TruncMonth("fecha_registro"))
            .values("mes")
            .annotate(total=Count("id"))
            .order_by("mes")
        )

        # Formatear mes legible
        result = []
        for item in data:
            result.append({
                "mes": item["mes"].strftime("%Y-%m"),
                "total": item["total"]
            })

        return Response(result)


# =============================
# ESTADÍSTICA: USUARIOS EN SESIÓN
# (últimos 30 minutos)
# =============================
class UsuariosEnSesionView(APIView):
    permission_classes = [IsAuthenticated, IsAdministrador]

    def get(self, request):
        desde = timezone.now() - timezone.timedelta(minutes=30)

        total = Usuario.objects.filter(
            last_login__gte=desde
        ).count()

        return Response({
            "total_en_sesion": total
        })


# =============================
# EXPORTAR USUARIOS A EXCEL
# =============================
class ExportUsuariosExcelView(APIView):
    permission_classes = [IsAuthenticated, IsAdministrador]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        fecha_reporte = now().strftime("%d/%m/%Y %H:%M")

        # =============================
        # TÍTULO
        # =============================
        ws.merge_cells("A1:H1")
        ws["A1"] = f"Reporte de Usuarios - Generado: {fecha_reporte}"
        ws["A1"].font = Font(bold=True)

        # =============================
        # ENCABEZADOS
        # =============================
        headers = [
            "ID",
            "Identificación",
            "Nombre",
            "Apellido",
            "Email",
            "Rol",
            "Estado",
            "Fecha Registro",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col)].width = 20

        # =============================
        # DATOS
        # =============================
        usuarios = Usuario.objects.all().order_by("-id")

        row = 4
        for u in usuarios:
            ws.cell(row=row, column=1, value=u.id)
            ws.cell(row=row, column=2, value=u.identificacion)
            ws.cell(row=row, column=3, value=u.nombre)
            ws.cell(row=row, column=4, value=u.apellido)
            ws.cell(row=row, column=5, value=u.email)
            ws.cell(row=row, column=6, value=u.rol)
            ws.cell(row=row, column=7, value="Activo" if u.is_active else "Inactivo")
            ws.cell(row=row, column=8, value=u.fecha_registro.strftime("%Y-%m-%d"))
            row += 1

        # =============================
        # RESPONSE
        # =============================
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_usuarios.xlsx"'
        wb.save(response)

        return response