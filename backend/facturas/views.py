# facturas/views.py
from rest_framework import viewsets, status, generics
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.pagination import PageNumberPagination
from django.http import FileResponse
from django.core.mail import EmailMessage
from django.conf import settings
from django.utils import timezone
from django.utils.dateparse import parse_date

from config.permissions import IsAdministrador
from .models import Factura
from pedidos.models import Pedido
from pedidos_detalles.models import PedidoDetalle
from .serializers import FacturaSerializer

from django.db.models import Q
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

import io
import os

# ReportLab (mantengo tu generador existente)
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
from reportlab.lib.units import cm

# ---------------------------
# PAGINACIÓN (10 por página)
# ---------------------------
class TenResultsPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"   # opcionalmente permitir cambiar
    max_page_size = 100

# ---------------------------
# GENERADOR PDF (tu función existente)
# ---------------------------
def generar_pdf_factura(factura):
    # (copié exactamente tu implementación original para mantener detalles)
    buffer = io.BytesIO()

    pdf = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=50,
    )

    # Metadata del PDF
    pdf.title = f"Factura Electrónica #{factura.id}"
    pdf.author = "El Gustador"
    pdf.subject = "Factura electrónica de venta"
    pdf.keywords = "Factura, El Gustador, Venta"

    pedido = factura.pedido
    detalles = PedidoDetalle.objects.filter(pedido=pedido)

    elements = []
    styles = getSampleStyleSheet()

    # =================================================================
    #                           COLORES
    # =================================================================
    COLOR_NARANJA = colors.HexColor("#FF6A00")
    COLOR_VERDE   = colors.HexColor("#1EBE4E")
    COLOR_AZUL    = colors.HexColor("#1F2B5B")
    COLOR_GRIS    = colors.HexColor("#F3F3F3")

    # =================================================================
    #                           ESTILOS
    # =================================================================
    titulo_style = ParagraphStyle(
        "TituloFactura",
        alignment=1,
        fontSize=22,
        textColor=COLOR_AZUL,
        leading=24,
        spaceAfter=20,
        fontName="Helvetica-Bold",
    )

    encabezado_style = ParagraphStyle(
        "EncabezadoEmpresa",
        alignment=1,
        fontSize=10,
        textColor=colors.black,
        leading=14,
        spaceAfter=0,
    )

    texto_normal = ParagraphStyle("TextoNormal", fontSize=10, leading=14)
    texto_bold   = ParagraphStyle("TextoBold", fontSize=10, leading=14, fontName="Helvetica-Bold")

    # =================================================================
    #             ENCABEZADO ESTÁTICO + LOGO (MEJORADO)
    # =================================================================

    # Línea decorativa superior
    elements.append(
        Table(
            [[""]],
            colWidths=[520],
            style=[
                ("BACKGROUND", (0, 0), (-1, -1), COLOR_NARANJA),
                ("LINEBELOW", (0, 0), (-1, -1), 2, COLOR_VERDE),
            ],
        )
    )
    elements.append(Spacer(1, 10))

    encabezado_texto = """
        <b>EL GUSTADOR</b><br/>
        María Osnid Díaz Castillo<br/>
        NIT. 00000000000000<br/>
        Responsable de IVA<br/>
        Carrera 5 #145-75<br/>
        Ibagué - Tolima<br/>
        Tel: 3101010101<br/>
    """

    # Logo
    logo_path = os.path.join(settings.BASE_DIR, "media/empresa/logo.png")
    if os.path.exists(logo_path):
        logo_img = Image(logo_path, width=120, height=120)
    else:
        logo_img = Paragraph("", texto_normal)

    encabezado_tabla = Table(
        [
            [
                logo_img,
                Paragraph(encabezado_texto, encabezado_style),
            ]
        ],
        colWidths=[160, 360],
    )

    encabezado_tabla.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, 0), "CENTER"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    elements.append(encabezado_tabla)
    elements.append(Spacer(1, 20))

    # =================================================================
    #                           TÍTULO
    # =================================================================
    elements.append(Paragraph("Factura Electrónica", titulo_style))
    elements.append(Spacer(1, 14))

    # =================================================================
    #         INFORMACIÓN CLIENTE + INFORMACIÓN FACTURA
    # =================================================================

    info_cliente = f"""
        <b>INFORMACIÓN DEL CLIENTE</b><br/><br/>
        <b>Nombre:</b> {pedido.usuario.nombre}<br/>
        <b>Documento:</b> {pedido.usuario.identificacion}<br/>
        <b>Email:</b> {pedido.usuario.email}<br/>
        <b>Dirección:</b> {pedido.municipio.nombre} - {pedido.departamento.nombre} - {pedido.direccion_detallada}<br/>
    """

    info_factura = f"""
        <b>Factura #:</b> {factura.id}<br/>
        <b>Fecha:</b> {factura.fecha.strftime('%Y-%m-%d %H:%M')}<br/>
        <b>Estado Pedido:</b> {pedido.estado}<br/>
        <b>Pago:</b> Confirmado<br/>
    """

    tabla_info = Table(
        [[Paragraph(info_cliente, texto_normal),
          Paragraph(info_factura, texto_normal)]],
        colWidths=[260, 260]
    )

    tabla_info.setStyle(
        TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ])
    )

    elements.append(tabla_info)
    elements.append(Spacer(1, 20))

    # =================================================================
    #                        TABLA PRODUCTOS
    # =================================================================
    table_data = [["ARTÍCULO", "CANT", "PRECIO", "SUBTOTAL"]]

    for item in detalles:
        table_data.append([
            item.producto.nombre,
            str(item.cantidad),
            f"${item.precio_unitario:,.0f}",
            f"${item.precio_total:,.0f}",
        ])

    tabla_productos = Table(table_data, colWidths=[260, 80, 90, 90])

    tabla_productos.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), COLOR_NARANJA),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, COLOR_GRIS]),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )

    elements.append(tabla_productos)
    elements.append(Spacer(1, 20))

    # =================================================================
    #                             TOTALES
    # =================================================================
    total_items = sum(item.precio_total for item in detalles)

    tabla_totales = Table([
        ["Subtotal:", f"${total_items:,.0f}"],
        ["Envío:", f"${pedido.costo_envio:,.0f}"],
        ["TOTAL:", f"${factura.total:,.0f}"],
    ], colWidths=[400, 120])

    tabla_totales.setStyle(
        TableStyle([
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 2), (-1, 2), COLOR_VERDE),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
        ])
    )

    elements.append(tabla_totales)
    elements.append(Spacer(1, 25))

    # =================================================================
    #                           PIE DE PÁGINA 
    # =================================================================

    footer_line = Table(
        [[""]],
        colWidths=[520],
        style=[("LINEABOVE", (0, 0), (-1, -1), 1.2, COLOR_VERDE)]
    )
    elements.append(footer_line)
    elements.append(Spacer(1, 6))

    footer_text = """
        <para align='center'>
            <font color='#1EBE4E'><b>Gracias por tu compra</b></font><br/>
            <font color='#FF6A00'>www.elgustador.com</font>
        </para>
    """

    elements.append(
        Paragraph(footer_text, ParagraphStyle("footer", alignment=1, fontSize=11))
    )

    pdf.build(elements)
    buffer.seek(0)
    return buffer

# =================================================================
#                      VIEWSET PRINCIPAL (con filtros y paginación)
# =================================================================
class FacturaViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = FacturaSerializer
    pagination_class = TenResultsPagination

    def get_queryset(self):
        user = self.request.user
        qs = Factura.objects.all() if user.rol == "admin" else Factura.objects.filter(pedido__usuario=user)

        params = self.request.query_params

        # ID exacto (entero)
        factura_id = params.get("id")
        if factura_id:
            qs = qs.filter(id=factura_id)

        # Número factura parcial
        numero = params.get("numero_factura")
        if numero:
            qs = qs.filter(numero_factura__icontains=numero)

        # Buscador general
        q = params.get("q")
        if q:
            qs = qs.filter(
                Q(numero_factura__icontains=q) |
                Q(pedido__usuario__nombre__icontains=q)
            )

        # Filtros de fecha
        fecha_from = params.get("fecha_from")
        fecha_to = params.get("fecha_to")

        if fecha_from:
            d = parse_date(fecha_from)
            if d:
                qs = qs.filter(fecha__date__gte=d)

        if fecha_to:
            d2 = parse_date(fecha_to)
            if d2:
                qs = qs.filter(fecha__date__lte=d2)

        # Total mínimo/máximo
        total_min = params.get("total_min")
        total_max = params.get("total_max")

        if total_min:
            try:
                qs = qs.filter(total__gte=float(total_min))
            except:
                pass

        if total_max:
            try:
                qs = qs.filter(total__lte=float(total_max))
            except:
                pass

        return qs.order_by("-fecha")

    def list(self, request, *args, **kwargs):
        """
        `list` usa paginación automática (PageNumberPagination) y devuelve
        página de 10 resultados por defecto. Los params de filtro se pasan
        como query params.
        """
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True, context={"request": request})
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True, context={"request": request})
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        factura = serializer.save()
        return Response(self.get_serializer(factura).data, status=201)


# =================================================================
#                      LISTADOS (compatibles con filtros y paginación)
# =================================================================
class FacturasUsuarioView(generics.ListAPIView):
    serializer_class = FacturaSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = TenResultsPagination

    def get_queryset(self):
        user = self.request.user
        qs = Factura.objects.filter(pedido__usuario=user)

        params = self.request.query_params

        factura_id = params.get("id")
        if factura_id:
            qs = qs.filter(id=factura_id)

        q = params.get("q")
        if q:
            qs = qs.filter(
                Q(numero_factura__icontains=q) |
                Q(pedido__usuario__nombre__icontains=q)
            )

        fecha_from = params.get("fecha_from")
        fecha_to = params.get("fecha_to")

        if fecha_from:
            d = parse_date(fecha_from)
            if d:
                qs = qs.filter(fecha__date__gte=d)

        if fecha_to:
            d2 = parse_date(fecha_to)
            if d2:
                qs = qs.filter(fecha__date__lte=d2)

        total_min = params.get("total_min")
        total_max = params.get("total_max")

        if total_min:
            try:
                qs = qs.filter(total__gte=float(total_min))
            except:
                pass

        if total_max:
            try:
                qs = qs.filter(total__lte=float(total_max))
            except:
                pass

        return qs.order_by("-fecha")


class FacturasAdminView(generics.ListAPIView):
    serializer_class = FacturaSerializer
    permission_classes = [IsAuthenticated, IsAdministrador]
    pagination_class = TenResultsPagination

    def get_queryset(self):
        qs = Factura.objects.all()
        params = self.request.query_params

        # ID exacto
        factura_id = params.get("id")
        if factura_id:
            qs = qs.filter(id=factura_id)

        # Buscador general
        q = params.get("q")
        if q:
            qs = qs.filter(
                Q(numero_factura__icontains=q) |
                Q(pedido__usuario__nombre__icontains=q)
            )

        # Fechas
        fecha_from = params.get("fecha_from")
        fecha_to = params.get("fecha_to")

        if fecha_from:
            d = parse_date(fecha_from)
            if d:
                qs = qs.filter(fecha__date__gte=d)

        if fecha_to:
            d2 = parse_date(fecha_to)
            if d2:
                qs = qs.filter(fecha__date__lte=d2)

        # Totales
        total_min = params.get("total_min")
        total_max = params.get("total_max")

        if total_min:
            try:
                qs = qs.filter(total__gte=float(total_min))
            except:
                pass

        if total_max:
            try:
                qs = qs.filter(total__lte=float(total_max))
            except:
                pass

        return qs.order_by("-fecha")


# =================================================================
#                     DESCARGAR FACTURA PDF
# =================================================================
class DescargarFacturaPDF(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, factura_id):
        try:
            factura = Factura.objects.get(id=factura_id)
        except Factura.DoesNotExist:
            return Response({"detail": "Factura no encontrada"}, status=404)

        # Seguridad
        if factura.pedido.usuario != request.user and request.user.rol != "admin":
            return Response({"detail": "No tienes permiso"}, status=403)

        pdf_buffer = generar_pdf_factura(factura)

        return FileResponse(pdf_buffer, as_attachment=True, filename=f"factura_{factura.id}.pdf")


# =================================================================
#                     ENVIAR POR EMAIL
# =================================================================
class EnviarFacturaEmail(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, factura_id):
        try:
            factura = Factura.objects.get(id=factura_id)
        except Factura.DoesNotExist:
            return Response({"detail": "Factura no encontrada"}, status=404)

        # Seguridad
        if factura.pedido.usuario != request.user and request.user.rol != "admin":
            return Response({"detail": "No tienes permiso"}, status=403)

        # Generar PDF
        pdf_buffer = generar_pdf_factura(factura)
        pdf_bytes = pdf_buffer.getvalue()

        if not pdf_bytes or len(pdf_bytes) < 100:
            return Response({"detail": "Error generando el PDF"}, status=500)

        # ----------------------------------------------------------
        #                PLANTILLA DE CORREO (HTML)
        # ----------------------------------------------------------
        html_body = f"""
        <div style="font-family: Arial, sans-serif; padding: 20px; color:#333;">
            
            <div style="text-align: center; margin-bottom: 30px;">
                <h2 style="color:#1F2B5B; margin-bottom: 5px;">Factura Electrónica de Venta</h2>
                <p style="font-size: 14px; color:#555;">
                    Gracias por tu compra. Adjuntamos la factura correspondiente a tu pedido.
                </p>
            </div>

            <div style="background:#F9F9F9; padding: 15px; border-radius: 8px; border-left:5px solid #FF6A00;">
                <p><strong style="color:#1EBE4E;">Factura N.º:</strong> {factura.id}</p>
                <p><strong style="color:#1EBE4E;">Fecha:</strong> {factura.fecha.strftime('%Y-%m-%d %H:%M')}</p>
                <p><strong style="color:#1EBE4E;">Cliente:</strong> {factura.pedido.usuario.nombre}</p>
            </div>

            <p style="margin-top: 25px; line-height: 1.6;">
                Hemos generado tu factura en formato PDF.  
                Puedes descargarla en el archivo adjunto.
            </p>

            <hr style="border: none; border-top: 1px solid #EEE; margin: 30px 0;">

            <p style="text-align: center; font-size: 12px; color:#777;">
                © {timezone.now().year} El Gustador ·  
                <a href="https://elgustador.com" style="color:#FF6A00; text-decoration:none;">
                    elgustador.com
                </a><br>
                Gracias por confiar en nosotros.
            </p>
        </div>
        """

        # ----------------------------------------------------------
        #                      ENVÍO DEL CORREO
        # ----------------------------------------------------------
        email = EmailMessage(
            subject=f"Factura Electrónica de Venta #{factura.id}",
            body=html_body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[factura.pedido.usuario.email],
        )

        # enviar como HTML
        email.content_subtype = "html"

        # Adjuntar PDF
        email.attach(
            f"factura_{factura.id}.pdf",
            pdf_bytes,
            "application/pdf"
        )

        email.send()

        return Response({"detail": "Factura enviada correctamente al correo"}, status=200)

class FacturasAdminSearchView(generics.ListAPIView):
    permission_classes = [IsAuthenticated, IsAdministrador]
    serializer_class = FacturaSerializer
    pagination_class = TenResultsPagination

    def get_queryset(self):
        qs = Factura.objects.select_related(
            "pedido", "pedido__usuario"
        ).all()

        q = self.request.query_params.get("q")

        if q:
            qs = qs.filter(
                Q(id__icontains=q) |
                Q(numero_factura__icontains=q) |
                Q(pedido__id__icontains=q) |
                Q(pedido__usuario__nombre__icontains=q) |
                Q(pedido__usuario__identificacion__icontains=q)
            )

        return qs.order_by("-fecha")

class FacturasAdminExportExcel(APIView):
    permission_classes = [IsAuthenticated, IsAdministrador]

    def get(self, request):
        fecha_from = request.query_params.get("fecha_from")
        fecha_to = request.query_params.get("fecha_to")

        qs = Factura.objects.select_related(
            "pedido", "pedido__usuario"
        ).all()

        if fecha_from:
            qs = qs.filter(fecha__date__gte=parse_date(fecha_from))
        if fecha_to:
            qs = qs.filter(fecha__date__lte=parse_date(fecha_to))

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas"

        # Encabezados
        headers = [
            "Factura ID", "Número Factura", "Fecha",
            "Total", "Estado Pedido",
            "Pedido ID",
            "Usuario Nombre", "Usuario Identificación", "Usuario Email",
            "Producto", "Cantidad", "Precio Unitario", "Subtotal"
        ]
        ws.append(headers)

        for factura in qs:
            pedido = factura.pedido
            detalles = PedidoDetalle.objects.filter(pedido=pedido)

            # Si no hay detalles, al menos colocar información de factura
            if not detalles.exists():
                ws.append([
                    factura.id,
                    factura.numero_factura,
                    factura.fecha.strftime("%Y-%m-%d %H:%M"),
                    factura.total,
                    pedido.estado,
                    pedido.id,
                    pedido.usuario.nombre,
                    pedido.usuario.identificacion,
                    pedido.usuario.email,
                    "", "", "", ""
                ])
            else:
                for d in detalles:
                    ws.append([
                        factura.id,
                        factura.numero_factura,
                        factura.fecha.strftime("%Y-%m-%d %H:%M"),
                        factura.total,
                        pedido.estado,
                        pedido.id,
                        pedido.usuario.nombre,
                        pedido.usuario.identificacion,
                        pedido.usuario.email,
                        d.producto.nombre,
                        d.cantidad,
                        d.precio_unitario,
                        d.precio_total,
                    ])

        # Ajustar ancho automático
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Generar respuesta
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="facturas.xlsx"'
        wb.save(response)
        return response
