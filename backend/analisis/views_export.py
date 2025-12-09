# views_export.py
import io
from datetime import datetime
import pandas as pd
from django.http import HttpResponse
from django.apps import apps
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

class ExportModelExcelView(APIView):
    """
    Exporta cualquier modelo de la base de datos a Excel con estilo.
    Parámetros en la URL:
        ?model=Usuario
        ?app=usuarios
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        model_name = request.GET.get("model")
        app_name = request.GET.get("app")

        if not model_name or not app_name:
            return HttpResponse("Faltan parámetros 'model' y 'app'", status=400)

        # Obtener el modelo dinámicamente
        try:
            Model = apps.get_model(app_name, model_name)
        except LookupError:
            return HttpResponse("Modelo no encontrado", status=404)

        # Convertir queryset a DataFrame
        qs = Model.objects.all().values()
        df = pd.DataFrame.from_records(qs)

        # ===== Convertir datetimes timezone-aware a naive =====
        for col in df.select_dtypes(include=["datetimetz"]).columns:
            df[col] = df[col].dt.tz_localize(None)

        # Crear Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=model_name, startrow=2)
            workbook = writer.book
            ws = writer.sheets[model_name]

            # =================================
            # Encabezado principal
            # =================================
            header_text = f"{model_name} - Exportación"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            cell = ws.cell(row=1, column=1)
            cell.value = header_text
            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # naranja
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # =================================
            # Fecha de generación
            # =================================
            fecha_text = f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
            cell = ws.cell(row=2, column=1)
            cell.value = fecha_text
            cell.font = Font(italic=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # =================================
            # Ajustar ancho de columnas automáticamente
            # =================================
            for col_num, column in enumerate(df.columns, 1):
                max_length = max(df[column].astype(str).map(len).max(), len(column))
                ws.column_dimensions[ws.cell(row=3, column=col_num).column_letter].width = max_length + 5

            # =================================
            # Crear tabla de Excel (Control+T)
            # =================================
            tab = Table(
                displayName=f"Table_{model_name}",
                ref=f"A3:{ws.cell(row=df.shape[0]+3, column=df.shape[1]).coordinate}"
            )
            style = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

        # Enviar archivo
        output.seek(0)
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{model_name}.xlsx"'
        return response
